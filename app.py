import json
import os
from concurrent.futures import ThreadPoolExecutor
import time
import traceback
import uuid
from datetime import datetime, timezone
from io import BytesIO

import requests
from flask import Flask, request, jsonify, send_from_directory, send_file
from databricks.sdk import WorkspaceClient
from databricks.sdk.core import Config
from databricks.sdk.service.sql import StatementParameterListItem, StatementState

app = Flask(__name__, static_folder="static", static_url_path="")

def _require_env(name):
    v = (os.getenv(name) or "").strip()
    if not v:
        raise RuntimeError(
            f"Required environment variable {name!r} is not set. "
            f"Configure it in app.yaml (see README 'Deploy to your workspace')."
        )
    return v

# Required — no defaults, since these encode workspace-specific resources.
CATALOG = _require_env("CATALOG")
SCHEMA = _require_env("SCHEMA")
WAREHOUSE_ID = _require_env("DATABRICKS_WAREHOUSE_ID")
TABLE = f"{CATALOG}.{SCHEMA}.discovery"

# Optional — sensible defaults are OK here.
COE_GROUP = os.getenv("COE_GROUP_NAME") or "genie-coe-reviewers"
# BO group: members can view all engagements but only edit S1/S2 + click BO
# Approved on benchmark rows. If unset, no users get BO permissions.
BO_GROUP = os.getenv("BO_GROUP_NAME") or "genie-bo-reviewers"

# Optional. Microsoft Teams Incoming Webhook URL (Power Automate "Workflows"
# style, or a legacy O365 connector). When an analyst marks an engagement
# "Ready for COE Review", the app posts an Adaptive Card to this webhook's
# channel — which is how the COE group gets notified (everyone in that Teams
# channel sees it). If unset/empty, the notification is silently skipped and
# the status flip still succeeds. Best-effort only; never blocks the workflow.
TEAMS_COE_WEBHOOK_URL = (os.getenv("TEAMS_COE_WEBHOOK_URL") or "").strip()

w = WorkspaceClient()

# Section columns grouped by session (deterministic, no-AI v2 — 4 sessions).
#   S1 Business Context
#   S2 Key Terms & Metrics + Questions (each question carries a `type`:
#      benchmark | testing | out_of_scope | clarifying)
#   S3 Technical Design (slimmed): data sources + warehouse, global filter
#      comment, text instructions, data gaps, scope boundaries
#   S4 COE Review (approval gate only)
# `data_plan` (data sources) and `plan_warehouse_id` (warehouse) were relocated
# here from the old S4/S5. Columns from the removed AI sessions are intentionally
# NOT dropped from the Delta table (ensure_table is additive) — they just fall
# out of this map and stop being read/written.
SESSION_COLS = {
    1: ["business_context", "pain_points", "existing_reports"],
    2: ["question_bank", "vocabulary_metrics"],
    3: ["data_plan", "plan_warehouse_id", "text_instructions",
        "data_gaps", "scope_boundaries", "global_filter"],
    4: ["coe_approval_status", "coe_approval_notes", "coe_reviewer_email"],
}

# The highest session number. Used so save_session's progress-pointer cap
# isn't hardcoded to a magic number.
LAST_SESSION = max(SESSION_COLS)

# Columns that store plain strings (not JSON-encoded structured data).
SCALAR_COLS = {
    "global_filter",
    "plan_warehouse_id",
    "coe_approval_status", "coe_approval_notes", "coe_reviewer_email",
}

# Columns whose JSON shape is an object (not an array). None remain in v2
# (all were on the removed AI sessions); data_plan is an array.
OBJECT_COLS = set()


def _default_section_value(col):
    """The empty-default for a section column (string scalar, [] for arrays, {} for objects)."""
    if col in SCALAR_COLS:
        return ""
    if col in OBJECT_COLS:
        return "{}"
    return "[]"

# All section columns
ALL_SECTION_COLS = sorted(set(col for cols in SESSION_COLS.values() for col in cols))


# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------

class SqlTransientError(RuntimeError):
    """Raised when a SQL statement does not reach a successful terminal state
    within the warehouse-warmup window — typically a cold warehouse, network
    blip, or timeout. Distinct from a genuine empty result so callers can tell
    'row doesn't exist' from 'we couldn't ask.'"""


def sql_exec(query, params=None):
    """Execute a read query and return a list of dict rows.

    Returns:
        list[dict]: rows on a SUCCEEDED query (empty list = genuinely no rows).

    Raises:
        SqlTransientError: warehouse cold-start / timeout / non-terminal state.
        RuntimeError: query failed (syntax error, permission denied, etc).

    The wait_timeout is set generously so a cold warehouse has time to spin
    up before we treat the call as transient. Without this, every read
    against a cold warehouse silently returned [] and callers interpreted
    'no results' the same as 'real empty result' — for /api/jobs/start
    that surfaced as a confusing 404 'Engagement not found.'
    """
    sdk_params = None
    if params:
        sdk_params = [
            StatementParameterListItem(name=k, value=str(v) if v is not None else "")
            for k, v in params.items()
        ]
    resp = w.statement_execution.execute_statement(
        warehouse_id=WAREHOUSE_ID,
        statement=query,
        parameters=sdk_params,
        catalog=CATALOG,
        schema=SCHEMA,
        wait_timeout="50s",  # generous: covers warehouse cold-start
    )

    state = resp.status.state if resp.status else None
    if state == StatementState.SUCCEEDED:
        if resp.result and resp.result.data_array:
            cols = [c.name for c in resp.manifest.schema.columns]
            return [dict(zip(cols, row)) for row in resp.result.data_array]
        return []  # genuine empty result
    if state in (StatementState.FAILED, StatementState.CANCELED, StatementState.CLOSED):
        err = (
            resp.status.error.message
            if resp.status and resp.status.error
            else f"state={state}"
        )
        raise RuntimeError(f"SQL failed: {err}")
    # PENDING / RUNNING / None — non-terminal after wait_timeout
    raise SqlTransientError(
        f"SQL did not complete within wait_timeout (state={state}). "
        "The warehouse may be cold; retry in a few seconds."
    )


def sql_run(query, params=None, *, must_succeed=False):
    """Execute a write statement.

    By default this is fire-and-forget for back-compat -- ensure_table()'s
    CREATE TABLE IF NOT EXISTS calls would otherwise crash startup if the
    app SP lacks CREATE TABLE on the schema (the table may already exist;
    permission is checked before IF NOT EXISTS short-circuits).

    Pass must_succeed=True for writes where silent failure is unacceptable
    (engagement saves, soft-delete, plan persistence). That path waits for
    completion and raises on any non-SUCCEEDED state.
    """
    sdk_params = None
    if params:
        sdk_params = [
            StatementParameterListItem(name=k, value=str(v) if v is not None else "")
            for k, v in params.items()
        ]
    if not must_succeed:
        # Best-effort fire-and-forget. Wait briefly so quick statements
        # commit before we return, but don't raise on failure.
        w.statement_execution.execute_statement(
            warehouse_id=WAREHOUSE_ID,
            statement=query,
            parameters=sdk_params,
            catalog=CATALOG,
            schema=SCHEMA,
            wait_timeout="10s",
        )
        return

    resp = w.statement_execution.execute_statement(
        warehouse_id=WAREHOUSE_ID,
        statement=query,
        parameters=sdk_params,
        catalog=CATALOG,
        schema=SCHEMA,
        wait_timeout="30s",
    )
    state = str(resp.status.state) if resp.status else ""
    statement_id = resp.statement_id
    deadline = time.time() + 60
    while statement_id and ("PENDING" in state or "RUNNING" in state):
        if time.time() > deadline:
            raise RuntimeError(f"sql_run timed out after 90s: {query[:80]}...")
        time.sleep(1.0)
        resp = w.statement_execution.get_statement(statement_id)
        state = str(resp.status.state) if resp.status else ""
    if "SUCCEEDED" not in state:
        err = resp.status.error.message if (resp.status and resp.status.error) else state
        raise RuntimeError(f"sql_run failed [{state}]: {err}; query: {query[:200]}")


def now_ts():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def get_current_user():
    email = (
        request.headers.get("X-Forwarded-Email")
        or request.headers.get("X-Forwarded-User")
        or request.headers.get("X-Forwarded-Preferred-Username")
    )
    if email:
        return email
    try:
        return w.current_user.me().user_name
    except Exception:
        return "unknown"


def _empty_for_col(col):
    """In-memory empty for a column (after parse): str / list / dict."""
    if col in SCALAR_COLS:
        return ""
    if col in OBJECT_COLS:
        return {}
    return []


def _decode_section_value(col, raw):
    """Decode a stored JSON section column, falling back to the column's
    empty default if it's missing or unparseable."""
    if not raw:
        return _empty_for_col(col)
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return _empty_for_col(col)


def parse_row(row):
    """Parse a raw DB row: decode JSON section columns, build sessions dict."""
    eng = {}
    for k, v in row.items():
        if k in ALL_SECTION_COLS:
            if k in SCALAR_COLS:
                eng[k] = v or ""
            else:
                eng[k] = _decode_section_value(k, v)
        else:
            eng[k] = v

    eng["sessions"] = {}
    for snum, cols in SESSION_COLS.items():
        session = {}
        for c in cols:
            session[c] = eng.get(c, _empty_for_col(c))
        eng["sessions"][str(snum)] = session
    return eng


# ---------------------------------------------------------------------------
# DB migration: ensure new columns exist
# ---------------------------------------------------------------------------

def ensure_table():
    """Create the engagement Delta table on first run, then add any missing columns.

    The README promises auto-creation; this is that promise. Requires the SP to
    have CREATE TABLE on <CATALOG>.<SCHEMA>.
    """
    section_ddl = ", ".join(f"{col} STRING" for col in ALL_SECTION_COLS)
    sql_run(
        f"CREATE TABLE IF NOT EXISTS {TABLE} ("
        f"engagement_id STRING, genie_space_name STRING, "
        f"business_owner_name STRING, business_owner_email STRING, "
        f"analyst_name STRING, analyst_email STRING, "
        f"servicenow_ticket_url STRING, "
        f"current_session INT, status STRING, "
        f"created_at STRING, updated_at STRING, deleted_at STRING, "
        f"{section_ddl}"
        f") USING DELTA"
    )
    rows = sql_exec(f"DESCRIBE TABLE {TABLE}")
    existing = {r.get("col_name", "") for r in rows}
    # Top-level metadata columns added since v1
    for top_col in ("servicenow_ticket_url", "deleted_at"):
        if top_col not in existing:
            sql_run(f"ALTER TABLE {TABLE} ADD COLUMN {top_col} STRING")
    for col in ALL_SECTION_COLS:
        if col not in existing:
            sql_run(f"ALTER TABLE {TABLE} ADD COLUMN {col} STRING")

ensure_table()


# ---------------------------------------------------------------------------
# API: User
# ---------------------------------------------------------------------------

@app.route("/api/user")
def api_user():
    return jsonify({"email": get_current_user()})


def _user_workspace_client_from_token(user_token):
    """Build a WorkspaceClient from an explicit OBO token. Used by background
    job threads which can't access Flask's request context."""
    if not user_token:
        return None
    from databricks.sdk import WorkspaceClient as WC
    from databricks.sdk.core import Config
    cfg = Config(host=w.config.host, token=user_token, auth_type="pat")
    return WC(config=cfg)


def _user_workspace_client():
    """Build a WorkspaceClient using the forwarded user access token (OBO).
    Pulls the token from the current Flask request. Returns None if no token.
    """
    return _user_workspace_client_from_token(request.headers.get("X-Forwarded-Access-Token"))


def _user_is_coe_member(user_w):
    """True if the OBO user is in COE_GROUP. False on any error."""
    if not user_w:
        return False
    try:
        me = user_w.current_user.me()
        for g in (me.groups or []):
            if g.display == COE_GROUP:
                return True
    except Exception:
        pass
    return False


def _user_is_bo_member(user_w):
    """True if the OBO user is in BO_GROUP. False on any error."""
    if not user_w:
        return False
    try:
        me = user_w.current_user.me()
        for g in (me.groups or []):
            if g.display == BO_GROUP:
                return True
    except Exception:
        pass
    return False


def _user_role(user_w):
    """Return ('coe' | 'bo' | 'analyst' | None) — the user's role for permission
    decisions. COE wins over BO; default authenticated users are 'analyst'.
    Returns None only if there's no OBO client (unauthenticated request).
    """
    if not user_w:
        return None
    is_coe = _user_is_coe_member(user_w)
    if is_coe:
        return "coe"
    if _user_is_bo_member(user_w):
        return "bo"
    return "analyst"


def _authorize_engagement(eid):
    """Verify engagement exists. Read access is open to any authenticated user
    (CAN_USE on the app is the access boundary). Mutation gating happens at the
    section level via _gate_engagement_routes — see permissions model below.

    Permissions model:
      - Default (any authenticated user): full read/write on all engagements,
        EXCEPT clicking Approve in S4 (COE only) and clicking BO Approved on
        benchmark rows (COE or BO group only).
      - COE group: full access including Approve and BO Approved.
      - BO group: read all; edit S1/S2 only; view S4 (read-only except for BO
        Approved checkboxes); cannot touch S3/S5/S6 or any AI/push action.

    Returns (eng, None) on success, or (None, error_response) if not found
    (404) or the warehouse couldn't answer (503 — transient, retry).
    """
    try:
        rows = sql_exec(f"SELECT * FROM {TABLE} WHERE engagement_id = :eid", {"eid": eid})
    except SqlTransientError:
        # Cold warehouse / timeout — distinct from "row doesn't exist." Return
        # 503 so the frontend can show "wait and retry" instead of confusing
        # the user with "Not found."
        return None, (jsonify({
            "error": "Database temporarily unavailable. The SQL warehouse may "
                     "be starting up; retry in a few seconds.",
            "transient": True,
        }), 503)
    if not rows:
        return None, (jsonify({"error": "Engagement not found"}), 404)
    return parse_row(rows[0]), None


import re as _re

_UUID_RE = _re.compile(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$")


def _bo_can_access(method, sub_path):
    """Whitelist of operations a BO-only user (not COE) is allowed to perform.

    Everything else returns 403 for BOs. The whitelist is intentionally tight
    so adding new endpoints doesn't accidentally open BO access — if a future
    endpoint should be BO-accessible, it must be added here explicitly.
    """
    if method == "GET":
        return True
    if method == "PUT" and sub_path in ("/sessions/1", "/sessions/2"):
        return True
    # ServiceNow URL lives in the Session 1 view BOs can edit.
    if method == "PATCH" and sub_path == "/servicenow-url":
        return True
    # Pre-work Excel upload: BOs can upload their own filled-in template and
    # apply it to S1/S2 (the only sessions they have edit rights to anyway).
    # Export is read-only and covers the same S1/S2 data they can already see.
    if method == "POST" and sub_path in (
        "/parse-prework", "/apply-prework", "/export-prework"
    ):
        return True
    return False


@app.before_request
def _gate_engagement_routes():
    """Apply _authorize_engagement to every /api/engagements/<eid>[/...] route,
    then layer BO restrictions on top.

    Layer 1: existence/auth — engagement exists and caller is authenticated.
    Layer 2: section-level role gate — BO-only users are limited to a tight
             whitelist of operations (read everything, edit S1/S2, click the
             BO-Approved checkbox via the dedicated PATCH endpoint).

    Only triggers when the path segment after /api/engagements/ looks like an
    engagement UUID (so sibling routes like /api/engagements/check-name are
    not mistakenly gated).
    """
    path = request.path or ""
    prefix = "/api/engagements/"
    if not path.startswith(prefix):
        return None
    remainder = path[len(prefix):]
    if not remainder:
        return None
    eid = remainder.split("/", 1)[0]
    if not _UUID_RE.match(eid):
        return None
    _, err = _authorize_engagement(eid)
    if err:
        return err
    # Layer 2: BO-only users get a tight whitelist of operations.
    sub_path = remainder[len(eid):]  # "" or "/sessions/1" or "/coe-approve" etc.
    user_w = _user_workspace_client()
    role = _user_role(user_w)
    if role == "bo" and not _bo_can_access(request.method, sub_path):
        return jsonify({
            "error": f"Members of the '{BO_GROUP}' group cannot perform this action."
        }), 403
    return None


@app.route("/api/warehouses")
def api_warehouses():
    """List SQL warehouses visible to the current user (OBO).

    OBO-only so users only see warehouses they actually have access to. If the
    user's token is missing the `sql` scope (common after app scope changes
    until the user re-authorizes), return a 403 with an actionable message —
    do NOT fall back to the SP client, which would let users pick warehouses
    they can't execute against.
    """
    user_w = _user_workspace_client()
    if not user_w:
        return jsonify({"error": "No user access token"}), 401
    try:
        whs = list(user_w.warehouses.list())
    except Exception as e:
        tb = traceback.format_exc()
        print(f"[/api/warehouses] ERROR: {type(e).__name__}: {e}\n{tb}", flush=True)
        msg = str(e)
        if "does not have required scopes" in msg or "PermissionDenied" in type(e).__name__:
            return jsonify({
                "error": "Your app authorization is missing the `sql` scope. Please sign out of the app (or open in a private window) and re-authorize when prompted.",
                "reauth_required": True,
            }), 403
        return jsonify({"error": f"Failed to list warehouses: {type(e).__name__}: {e}"}), 500
    out = []
    for wh in whs:
        out.append({
            "id": wh.id,
            "name": wh.name,
            "state": str(wh.state) if wh.state else "",
            "size": wh.cluster_size or "",
            "type": str(wh.warehouse_type) if wh.warehouse_type else "",
        })
    out.sort(key=lambda x: x["name"].lower())
    return jsonify(out)


@app.route("/api/user/coe-member")
def api_user_coe_member():
    """Check COE membership using the user's forwarded access token (OBO).

    This respects the user's own permissions rather than requiring the app
    service principal to be a workspace admin. Append ?debug=1 for
    diagnostics.
    """
    email = get_current_user()
    debug = request.args.get("debug") == "1"
    result = {"is_member": False}
    if debug:
        result["email"] = email
        result["coe_group_name"] = COE_GROUP

    user_w = _user_workspace_client()
    if not user_w:
        if debug:
            result["error"] = "no X-Forwarded-Access-Token header"
        return jsonify(result)

    try:
        me = user_w.current_user.me()
        me_groups = me.groups or []
        if debug:
            result["me_group_count"] = len(me_groups)
            result["me_groups"] = [g.display for g in me_groups]
        for g in me_groups:
            if g.display == COE_GROUP:
                result["is_member"] = True
                return jsonify(result)
        return jsonify(result)
    except Exception as e:
        if debug:
            result["error"] = str(e)
            result["error_type"] = type(e).__name__
        return jsonify(result)


@app.route("/api/user/role")
def api_user_role():
    """Return the caller's role flags so the frontend can render the right UI.

    Frontend uses this to decide what tabs/buttons/checkboxes to show. The
    backend independently re-checks role on every mutation, so this is
    advisory-only — a frontend that fakes a role still gets 403'd server-side.
    """
    user_w = _user_workspace_client()
    return jsonify({
        "is_coe": _user_is_coe_member(user_w),
        "is_bo": _user_is_bo_member(user_w),
        "coe_group_name": COE_GROUP,
        "bo_group_name": BO_GROUP,
    })


# ---------------------------------------------------------------------------
# API: Engagements CRUD
# ---------------------------------------------------------------------------

@app.route("/api/engagements", methods=["GET"])
def list_engagements():
    """Return all non-deleted engagements. Read visibility is open to any
    authenticated user; the per-section write rules are enforced elsewhere
    (see permissions model in _authorize_engagement).
    """
    # Soft-deleted rows (deleted_at IS NOT NULL/empty) are hidden from the list.
    # The row stays in Delta so an admin can recover it via delete_at IS NULL/SQL.
    not_deleted = "(deleted_at IS NULL OR deleted_at = '')"
    try:
        rows = sql_exec(
            f"SELECT engagement_id, genie_space_name, business_owner_name, "
            f"business_owner_email, analyst_name, analyst_email, "
            f"current_session, status, coe_approval_status, created_at, updated_at "
            f"FROM {TABLE} WHERE {not_deleted} ORDER BY updated_at DESC"
        )
    except SqlTransientError:
        return jsonify({
            "error": "Database temporarily unavailable. The SQL warehouse may "
                     "be starting up; retry in a few seconds.",
            "transient": True,
        }), 503
    return jsonify(rows)


@app.route("/api/engagements/check-name")
def check_engagement_name():
    """Check if an engagement name is already taken.

    Optional `exclude_eid` query param: when renaming an existing engagement,
    its own current name should not count as a conflict.
    Soft-deleted engagements DO NOT count as conflicts -- the name is freed
    for reuse once the row is deleted.
    """
    name = request.args.get("name", "").strip()
    exclude_eid = request.args.get("exclude_eid", "").strip()
    if not name:
        return jsonify({"available": False})
    not_deleted = "(deleted_at IS NULL OR deleted_at = '')"
    if exclude_eid and _UUID_RE.match(exclude_eid):
        rows = sql_exec(
            f"SELECT COUNT(*) AS cnt FROM {TABLE} "
            f"WHERE genie_space_name = :name AND engagement_id != :eid "
            f"AND {not_deleted}",
            {"name": name, "eid": exclude_eid},
        )
    else:
        rows = sql_exec(
            f"SELECT COUNT(*) AS cnt FROM {TABLE} "
            f"WHERE genie_space_name = :name AND {not_deleted}",
            {"name": name},
        )
    count = int(rows[0]["cnt"]) if rows else 0
    return jsonify({"available": count == 0})


@app.route("/api/engagements", methods=["POST"])
def create_engagement():
    # BO-only users cannot create engagements (analysts and COE only).
    user_w = _user_workspace_client()
    if _user_role(user_w) == "bo":
        return jsonify({
            "error": f"Members of the '{BO_GROUP}' group cannot create engagements.",
        }), 403

    data = request.json
    # Validate required fields
    missing = []
    for field in ["genie_space_name", "business_owner_name", "business_owner_email", "analyst_name"]:
        if not data.get(field, "").strip():
            missing.append(field)
    if missing:
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

    # Check uniqueness
    name = data["genie_space_name"].strip()
    existing = sql_exec(
        f"SELECT COUNT(*) AS cnt FROM {TABLE} "
        f"WHERE genie_space_name = :name "
        f"AND (deleted_at IS NULL OR deleted_at = '')",
        {"name": name},
    )
    if existing and int(existing[0]["cnt"]) > 0:
        return jsonify({"error": "An engagement with this name already exists"}), 409

    eid = str(uuid.uuid4())
    ts = now_ts()
    section_cols_sql = ", ".join(ALL_SECTION_COLS)
    section_defaults = []
    section_params = {}
    for col in ALL_SECTION_COLS:
        param_name = f"default_{col}"
        section_defaults.append(f":{param_name}")
        section_params[param_name] = _default_section_value(col)

    sql_run(
        f"INSERT INTO {TABLE} "
        f"(engagement_id, genie_space_name, business_owner_name, business_owner_email, "
        f"analyst_name, analyst_email, servicenow_ticket_url, "
        f"current_session, status, created_at, updated_at, "
        f"{section_cols_sql}) "
        f"VALUES (:eid, :space_name, :bo_name, :bo_email, :a_name, :a_email, :sn_url, "
        f"1, 'draft', :ts, :ts, {', '.join(section_defaults)})",
        {
            "eid": eid,
            "space_name": name,
            "bo_name": data.get("business_owner_name", "").strip(),
            "bo_email": data.get("business_owner_email", "").strip(),
            "a_name": data.get("analyst_name", "").strip(),
            "a_email": data.get("analyst_email", "").strip(),
            "sn_url": (data.get("servicenow_ticket_url") or "").strip(),
            "ts": ts,
            **section_params,
        },
    )
    return jsonify({"engagement_id": eid}), 201


@app.route("/api/engagements/<eid>", methods=["GET"])
def get_engagement(eid):
    # SELECT * returns updated_at / created_at as TIMESTAMP, but we hand the
    # value out to clients as the optimistic-lock token. Cast to STRING so
    # the load value matches what _check_optimistic_lock will compare against
    # later (Spark's canonical TIMESTAMP-string format).
    rows = sql_exec(
        f"SELECT *, "
        f"CAST(updated_at AS STRING) AS updated_at_str, "
        f"CAST(created_at AS STRING) AS created_at_str "
        f"FROM {TABLE} WHERE engagement_id = :eid",
        {"eid": eid},
    )
    if not rows:
        return jsonify({"error": "Not found"}), 404
    parsed = parse_row(rows[0])
    # Overwrite the TIMESTAMP-typed fields with their string aliases
    if "updated_at_str" in rows[0]:
        parsed["updated_at"] = rows[0].get("updated_at_str") or ""
    if "created_at_str" in rows[0]:
        parsed["created_at"] = rows[0].get("created_at_str") or ""
    return jsonify(parsed)


@app.route("/api/engagements/<eid>", methods=["PUT"])
def update_engagement(eid):
    """Update engagement metadata. Validates genie_space_name uniqueness against
    every OTHER engagement (the current row's own name does not count as a
    conflict, so saving without renaming is always fine).
    """
    data = request.json or {}
    name = (data.get("genie_space_name") or "").strip()
    if not name:
        return jsonify({"error": "genie_space_name is required"}), 400

    # Optimistic-lock check (no-op if If-Match header is absent)
    try:
        _check_optimistic_lock(eid)
    except StaleEngagementError as e:
        return jsonify({
            "error": "stale",
            "current_updated_at": e.current_updated_at,
            "message": "This engagement was updated by another user. Refresh to continue.",
        }), 409
    except ValueError as e:
        return jsonify({"error": str(e)}), 404

    # Uniqueness check, scoped to other (non-soft-deleted) engagements
    dup = sql_exec(
        f"SELECT COUNT(*) AS cnt FROM {TABLE} "
        f"WHERE genie_space_name = :name AND engagement_id != :eid "
        f"AND (deleted_at IS NULL OR deleted_at = '')",
        {"name": name, "eid": eid},
    )
    if dup and int(dup[0]["cnt"]) > 0:
        return jsonify({"error": "An engagement with this name already exists"}), 409

    ts = now_ts()
    sql_run(
        f"UPDATE {TABLE} SET "
        f"genie_space_name = :space_name, business_owner_name = :bo_name, "
        f"business_owner_email = :bo_email, analyst_name = :a_name, "
        f"analyst_email = :a_email, servicenow_ticket_url = :sn_url, "
        f"status = :status, updated_at = :ts "
        f"WHERE engagement_id = :eid",
        {
            "eid": eid,
            "space_name": name,
            "bo_name": (data.get("business_owner_name") or "").strip(),
            "bo_email": (data.get("business_owner_email") or "").strip(),
            "a_name": (data.get("analyst_name") or "").strip(),
            "a_email": (data.get("analyst_email") or "").strip(),
            "sn_url": (data.get("servicenow_ticket_url") or "").strip(),
            "status": data.get("status", "draft"),
            "ts": ts,
        },
    )
    return jsonify({"success": True, "updated_at": _read_updated_at(eid)})


@app.route("/api/engagements/<eid>/servicenow-url", methods=["PATCH"])
def patch_servicenow_url(eid):
    """Update ONLY servicenow_ticket_url.

    Lightweight side-field write, modeled on the BO-Approved PATCH: it lives
    OUTSIDE the optimistic lock and does NOT bump updated_at, so it can never
    race/409 the analyst's session autosave (which would otherwise reload and
    silently revert the typed URL). It also doesn't touch current_session, so
    saving the URL never regresses engagement progress. Existence/auth is
    handled by the before_request gate.
    """
    data = request.json or {}
    url = (data.get("servicenow_ticket_url") or "").strip()
    sql_run(
        f"UPDATE {TABLE} SET servicenow_ticket_url = :u WHERE engagement_id = :eid",
        {"eid": eid, "u": url},
    )
    return jsonify({"success": True})


def _notify_teams_review_ready(eng, eid, analyst_email):
    """Post an Adaptive Card to the COE Teams channel when an engagement is
    marked Ready for COE Review. Best-effort: any failure (unset webhook,
    network error, bad response) is swallowed so it never blocks the status
    flip. Targets a channel webhook, so the whole COE group sees it.
    """
    if not TEAMS_COE_WEBHOOK_URL:
        return  # Feature not configured — silently skip.
    try:
        space_name = (eng.get("genie_space_name") or "Untitled Space") if eng else "Untitled Space"
        sn_url = (eng.get("servicenow_ticket_url") or "") if eng else ""
        # Deep link back to Section 4 of this engagement, derived from the
        # incoming request's host so we don't need a hardcoded app URL.
        try:
            base = request.host_url.rstrip("/")
        except Exception:
            base = ""
        eng_link = f"{base}/engagements/{eid}" if base else ""

        facts = [
            {"title": "Engagement", "value": space_name},
            {"title": "Marked ready by", "value": analyst_email or "an analyst"},
        ]
        if sn_url:
            facts.append({"title": "ServiceNow", "value": f"[Open ticket]({sn_url})"})

        body = [
            {
                "type": "TextBlock",
                "size": "Medium",
                "weight": "Bolder",
                "text": "🔔 Ready for COE Review",
            },
            {
                "type": "TextBlock",
                "text": f"**{space_name}** has been marked ready for Center of Excellence review.",
                "wrap": True,
            },
            {"type": "FactSet", "facts": facts},
        ]
        actions = []
        if eng_link:
            actions.append({
                "type": "Action.OpenUrl",
                "title": "Open in Genie Discovery",
                "url": eng_link,
            })

        card = {
            "type": "AdaptiveCard",
            "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
            "version": "1.4",
            "body": body,
        }
        if actions:
            card["actions"] = actions

        payload = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "content": card,
                }
            ],
        }
        requests.post(TEAMS_COE_WEBHOOK_URL, json=payload, timeout=10)
    except Exception as e:  # noqa: BLE001 — notification must never break the flip
        print(f"[teams] review-ready notification failed (non-fatal): {e}")


@app.route("/api/engagements/<eid>/request-review", methods=["PUT"])
def request_coe_review(eid):
    """Analyst action: mark an engagement 'Ready for COE Review'.

    Unlike coe-approve (COE-only), this is open to analysts — it only sets the
    status to 'ready_for_review', never to 'approved'/'changes_requested'
    (those stay COE-gated). BO-only users are blocked by the before_request
    gate since this path is NOT in _bo_can_access.

    Lightweight side-write modeled on the ServiceNow-URL PATCH: it lives
    OUTSIDE the optimistic lock and does NOT bump updated_at, so it can't
    race/409 the analyst's in-flight Session 4 autosave. The client mirrors
    the new value into its session-4 draft so the next autosave stays
    consistent. Fires the Teams notification best-effort on success.
    """
    eng, err = _authorize_engagement(eid)
    if err:
        return err
    # Marking ready re-enters the review flow, so a previously-completed
    # (Section 7 signed-off) engagement is no longer complete: roll status back
    # to 'in_progress' so it doesn't read as Complete on the home list. status
    # gates nothing; this keeps display consistent across surfaces.
    sql_run(
        f"UPDATE {TABLE} SET coe_approval_status = 'ready_for_review', "
        f"status = 'in_progress' "
        f"WHERE engagement_id = :eid",
        {"eid": eid},
    )
    _notify_teams_review_ready(eng, eid, get_current_user())
    return jsonify({"success": True, "status": "ready_for_review"})


@app.route("/api/engagements/<eid>", methods=["DELETE"])
def delete_engagement(eid):
    """Soft-delete: mark deleted_at instead of hard-removing the row.
    The row stays in Delta so a careless click doesn't lose months of work.
    Hidden from list_engagements + name-uniqueness checks; recoverable via
    direct URL or by clearing deleted_at via SQL.
    """
    ts = now_ts()
    sql_run(
        f"UPDATE {TABLE} SET deleted_at = :ts, updated_at = :ts "
        f"WHERE engagement_id = :eid",
        {"eid": eid, "ts": ts},
    )
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API: Session saves
# ---------------------------------------------------------------------------

class StaleEngagementError(Exception):
    """Raised when an optimistic-lock check on engagement.updated_at fails.
    The HTTP layer translates this into 409 with the current updated_at so
    the client can refresh + retry."""
    def __init__(self, current_updated_at):
        super().__init__("stale engagement")
        self.current_updated_at = current_updated_at


def _check_optimistic_lock(eid):
    """If the request carries `If-Match: <updated_at>`, verify it matches
    the row's current updated_at. Raises StaleEngagementError on conflict.
    No header = no check (back-compat).

    Reads updated_at via CAST AS STRING so the comparison uses Spark's
    canonical TIMESTAMP-string format (the column is TIMESTAMP, not STRING,
    so a raw read can come back in different shapes).
    """
    expected = (request.headers.get("If-Match") or "").strip()
    if not expected:
        return
    actual = _read_updated_at(eid)
    if not actual:
        raise ValueError("Engagement not found")
    if actual != expected:
        raise StaleEngagementError(actual)


def _read_updated_at(eid):
    """Read engagement.updated_at AS STRING so we get the same canonical
    Spark-formatted value we'd compare against later (avoids round-trip
    format mismatches between Python's isoformat and Spark's TIMESTAMP
    string display)."""
    rows = sql_exec(
        f"SELECT CAST(updated_at AS STRING) AS updated_at "
        f"FROM {TABLE} WHERE engagement_id = :eid",
        {"eid": eid},
    )
    if not rows:
        return ""
    return (rows[0].get("updated_at") or "").strip()


def save_session(eid, session_num, data):
    """Update session columns for an engagement. Returns the new updated_at
    timestamp (in Spark's canonical string format) on success. Raises
    StaleEngagementError if the request's If-Match header doesn't match
    the row's current updated_at."""
    _check_optimistic_lock(eid)

    cols = SESSION_COLS[session_num]
    set_parts = []
    ts = now_ts()
    params = {"eid": eid, "ts": ts}

    for col in cols:
        set_parts.append(f"{col} = :{col}")
        if col in SCALAR_COLS:
            params[col] = data.get(col, "")
        else:
            # Structured columns are JSON arrays (no object-shaped columns remain).
            default = {} if col in OBJECT_COLS else []
            params[col] = json.dumps(data.get(col, default))

    # Advance the progress pointer (capped at the final session). The terminal
    # state is owned exclusively by the S4 COE approval (see coe_approve):
    # saving a session NEVER changes a terminal status ('ready_for_pilot', or a
    # legacy 'complete'), it only holds/returns non-terminal engagements to
    # 'in_progress'.
    next_session = min(session_num + 1, LAST_SESSION)
    set_parts.append(f"current_session = GREATEST(current_session, {next_session})")
    set_parts.append(
        "status = CASE WHEN status IN ('complete', 'ready_for_pilot') "
        "THEN status ELSE 'in_progress' END"
    )

    set_parts.append("updated_at = :ts")
    set_sql = ", ".join(set_parts)

    sql_run(f"UPDATE {TABLE} SET {set_sql} WHERE engagement_id = :eid", params)
    # Return the canonical stored value -- not Python's isoformat -- so the
    # client's next If-Match round-trips correctly. updated_at is a TIMESTAMP
    # column and Spark reformats on read.
    return _read_updated_at(eid)


def _save_session_response(eid, session_num):
    """Shared wrapper for the per-session save routes: handles optimistic-lock
    conflicts and returns the new updated_at so the client can carry it forward."""
    try:
        ts = save_session(eid, session_num, request.json)
    except StaleEngagementError as e:
        return jsonify({
            "error": "stale",
            "current_updated_at": e.current_updated_at,
            "message": "This engagement was updated by another user. Refresh to continue.",
        }), 409
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    return jsonify({"success": True, "updated_at": ts})


@app.route("/api/engagements/<eid>/sessions/1", methods=["PUT"])
def save_session_1(eid):
    return _save_session_response(eid, 1)


@app.route("/api/engagements/<eid>/sessions/2", methods=["PUT"])
def save_session_2(eid):
    return _save_session_response(eid, 2)


@app.route("/api/engagements/<eid>/sessions/3", methods=["PUT"])
def save_session_3(eid):
    return _save_session_response(eid, 3)


@app.route("/api/engagements/<eid>/sessions/4", methods=["PUT"])
def save_session_4(eid):
    return _save_session_response(eid, 4)


# ---------------------------------------------------------------------------
# API: Business Owner Pre-Work Excel Upload
# ---------------------------------------------------------------------------
# Lets the analyst (or BO) send the BO a .xlsx template before the working
# session, then upload the filled-in version to populate Sessions 1 and 2.
#
# Three endpoints:
#   GET  /api/template/business-owner-prework.xlsx -- download template
#   POST /api/engagements/<eid>/parse-prework      -- parse + validate, NO mutation
#   POST /api/engagements/<eid>/apply-prework      -- atomic write, optimistic lock
#
# Apply semantics: each chosen section is REPLACED (not appended). The preview
# UI shows the diff so the user opts in section-by-section. This makes
# double-upload idempotent.

PREWORK_TEMPLATE_VERSION = "2.0"
PREWORK_MAX_BYTES = 5 * 1024 * 1024  # 5 MB

# Question Type values; must match Session2Form.tsx's Type select. Every
# question in the S2 Question Bank is flagged as one of these.
QUESTION_TYPE_OPTIONS = ["Benchmark", "Testing", "Out of scope", "Clarifying"]

# Standard S1 business-context questions. Pre-populated in the template (locked
# columns) so BOs only write the Notes column. Must stay in sync with
# Session1Form.tsx's CONTEXT_QUESTIONS.
_PREWORK_S1_CONTEXT = [
    ("What does your team do day-to-day?", "Scopes the question universe"),
    ("What decisions do you make with data?", "Identifies the high-value questions"),
    ("Who else on your team would use this?", "Sizes the audience and skill range"),
    ("How do you get ad hoc answers today?", "Reveals the bottleneck Genie solves"),
]

# Frequency dropdown values; must match Session1Form.tsx's REPORT_COLS select.
_PREWORK_FREQ_OPTIONS = ["Daily", "Weekly", "Monthly", "Quarterly", "Ad hoc"]

# Per-sheet config drives template generation AND parsing -- single source of
# truth so the two can't drift. Keys map to session column names.
_PREWORK_SHEETS = [
    {
        "name": "S1 Business Context",
        "key": "business_context",
        "session": 1,
        "headers": ["Question", "Why It Matters", "Your Notes"],
        "row_keys": ["question", "why_it_matters", "response"],
        "instruction": ("Answer each question in the 'Your Notes' column. "
                        "Don't edit the Question or Why It Matters columns."),
    },
    {
        "name": "S1 Pain Points",
        "key": "pain_points",
        "session": 1,
        "headers": ["Pain Point"],
        "row_keys": ["description"],
        "instruction": ("List the top frustrations your team has with getting "
                        "data answers today. One pain point per row."),
    },
    {
        "name": "S1 Existing Reports",
        "key": "existing_reports",
        "session": 1,
        "headers": ["Report/Dashboard Name", "What It Shows", "How Often Used", "Known Issues"],
        "row_keys": ["report_name", "what_it_shows", "frequency", "known_issues"],
        "instruction": ("Every report, dashboard, or spreadsheet your team "
                        "references regularly. 'How Often Used' must be one of: "
                        "Daily, Weekly, Monthly, Quarterly, Ad hoc."),
    },
    {
        "name": "S2 Question Bank",
        "key": "question_bank",
        "session": 2,
        "headers": ["Question", "Type", "Decision It Drives", "Clarification (Genie asks)"],
        "row_keys": ["question_text", "type", "decision_it_drives", "clarification"],
        "instruction": ("Real questions your team needs answered. Flag each "
                        "question's Type (Benchmark, Testing, Out of scope, or "
                        "Clarifying), note the decision it helps you make, and for "
                        "Clarifying questions fill the follow-up Genie should ask."),
    },
    {
        # Renamed from "Vocabulary & Metrics" in the app -- BOs were reading
        # the column labels as a pure glossary and missing that metrics belong
        # here too. Headers and seeded examples make the dual purpose explicit.
        "name": "S2 Key Terms & Metrics",
        "key": "vocabulary_metrics",
        "session": 2,
        "headers": ["Business Term or Metric",
                    "Definition or How It's Calculated",
                    "Other Names / Synonyms"],
        "row_keys": ["business_term", "what_they_mean", "synonyms"],
        "instruction": ("Include BOTH vocabulary (jargon, abbreviations, "
                        "filter logic) AND metrics (anything with a "
                        "calculation). If it's a number you report on, it "
                        "goes here. See the example rows below for both types."),
    },
]


def _build_prework_template():
    """Generate the pre-work .xlsx in memory and return a BytesIO.

    Pre-populates S1 Business Context with the standard prompts and seeds the
    Key Terms & Metrics sheet with three examples (one metric, one vocabulary
    term, one borderline case) to demonstrate the dual purpose. A hidden
    `_meta` sheet stores the template version so we can detect outdated
    uploads at parse time.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.column_dimensions["A"].width = 110
    ws_inst["A1"] = "Genie Discovery — Business Owner Pre-Work"
    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst["A3"] = ("Your analyst will load this file into the discovery app "
                     "to populate your engagement before the working session.")
    overview_lines = [
        "",
        "How to use this workbook:",
        "  1. Fill out each sheet listed in the tabs below.",
        "  2. Do NOT rename sheets or column headers — the app uses those names to find your answers.",
        "  3. Leave a row blank if you don't have content for it. Blank rows are skipped.",
        "  4. Save the file and send it back to your analyst.",
        "",
        "Sheets in this workbook:",
    ]
    for i, line in enumerate(overview_lines):
        ws_inst[f"A{4+i}"] = line
    overview_end = 4 + len(overview_lines)
    for i, sheet in enumerate(_PREWORK_SHEETS):
        ws_inst[f"A{overview_end+i}"] = f"  • {sheet['name']}: {sheet['instruction']}"
        ws_inst[f"A{overview_end+i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_inst.row_dimensions[overview_end + i].height = 45

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    instruction_font = Font(italic=True, color="555555")

    for sheet in _PREWORK_SHEETS:
        ws = wb.create_sheet(sheet["name"])
        # Row 1: instruction banner spanning all columns
        n_cols = len(sheet["headers"])
        ws.cell(row=1, column=1, value=sheet["instruction"]).font = instruction_font
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = 45
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        # Row 2: header
        for col_idx, header in enumerate(sheet["headers"], start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 40

        # Pre-fill standard S1 context questions on rows 3+
        if sheet["key"] == "business_context":
            for r_idx, (q, why) in enumerate(_PREWORK_S1_CONTEXT, start=3):
                ws.cell(row=r_idx, column=1, value=q)
                ws.cell(row=r_idx, column=2, value=why)
                # Wrap the static columns so long prompts are readable
                ws.cell(row=r_idx, column=1).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=r_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r_idx].height = 40

        # Frequency dropdown on existing_reports' column C
        if sheet["key"] == "existing_reports":
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(_PREWORK_FREQ_OPTIONS) + '"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add("C3:C200")

        # Type dropdown on question_bank's column B
        if sheet["key"] == "question_bank":
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(QUESTION_TYPE_OPTIONS) + '"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add("B3:B200")

        # Seed three question examples so the Type dropdown is visible in a
        # populated cell (BOs otherwise miss the click-to-reveal dropdown) and
        # the Clarification column's purpose is obvious. Parser skips these.
        if sheet["key"] == "question_bank":
            ws.cell(row=3, column=1, value="↓ EXAMPLES (delete these rows and replace with your own) ↓").font = (
                Font(italic=True, color="888888", bold=True)
            )
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
            q_examples = [
                ("What was our ED volume by site last month?", "Benchmark",
                 "Staffing / capacity decisions", ""),
                ("Do you mean inpatient or ED visits?", "Clarifying",
                 "", "Ask: inpatient or ED encounters?"),
                ("Can I see individual provider salaries?", "Out of scope",
                 "", ""),
            ]
            for r_idx, (q, typ, dec, clar) in enumerate(q_examples, start=4):
                for c_idx, val in enumerate((q, typ, dec, clar), start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = Font(italic=True, color="888888")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r_idx].height = 30

        # Seed three examples on the Key Terms & Metrics sheet to demonstrate
        # that BOTH metrics and vocabulary belong here. The parser detects and
        # skips these rows so they don't import as real data even if the BO
        # forgets to delete them.
        if sheet["key"] == "vocabulary_metrics":
            ws.cell(row=6, column=1, value="↓ EXAMPLES (delete these rows and replace with your own) ↓").font = (
                Font(italic=True, color="888888", bold=True)
            )
            ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=3)
            examples = [
                ("Net Revenue",
                 "Gross sales minus returns and discounts. This is a METRIC -- include calculation logic.",
                 "NR, Net Sales"),
                ("Active Customer",
                 "Customer with at least one order in the trailing 90 days. METRIC -- definition includes the rule.",
                 "Active account"),
                ("SKU",
                 "Stock-keeping unit; the unique identifier for a product. VOCABULARY -- just the definition.",
                 "Item ID, Product code"),
            ]
            for r_idx, (term, defn, syn) in enumerate(examples, start=7):
                ws.cell(row=r_idx, column=1, value=term).font = Font(italic=True, color="888888")
                ws.cell(row=r_idx, column=2, value=defn).font = Font(italic=True, color="888888")
                ws.cell(row=r_idx, column=3, value=syn).font = Font(italic=True, color="888888")
                ws.cell(row=r_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r_idx].height = 30

    # Hidden metadata sheet for version detection at parse time
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_version"
    meta["B1"] = PREWORK_TEMPLATE_VERSION
    meta["A2"] = "generated_at"
    meta["B2"] = now_ts()
    meta.sheet_state = "hidden"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_prework_export(selected_keys, data):
    """Build a pre-work .xlsx populated with the engagement's current S1/S2 data.

    Mirrors the blank template's sheet names, headers, and `_meta` version so an
    exported file can be edited and re-uploaded through parse/apply-prework
    (full round-trip). Only the sheets whose key is in `selected_keys` are
    included; `data` is {section_key: [row_dict, ...]} already normalized to the
    section's `row_keys` by the caller.

    Differs from _build_prework_template: rows are filled from `data`, the S1
    business-context prompts are NOT auto-seeded (they come in via the data
    rows), and the Key Terms example rows are NOT seeded.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # Preserve the canonical sheet order; include only the selected sheets.
    sheets = [s for s in _PREWORK_SHEETS if s["key"] in selected_keys]

    wb = Workbook()
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.column_dimensions["A"].width = 110
    ws_inst["A1"] = "Genie Discovery — Engagement Export"
    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst["A3"] = ("Exported from the discovery app with this engagement's "
                     "current answers. You can edit it and load it back via "
                     "'Upload Pre-Work' to apply changes to Sessions 1 & 2. "
                     "Don't rename sheets or column headers.")
    ws_inst["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_inst.row_dimensions[3].height = 45
    for i, sheet in enumerate(sheets):
        ws_inst[f"A{5+i}"] = f"  • {sheet['name']}: {sheet['instruction']}"
        ws_inst[f"A{5+i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_inst.row_dimensions[5 + i].height = 45

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    instruction_font = Font(italic=True, color="555555")

    for sheet in sheets:
        ws = wb.create_sheet(sheet["name"])
        n_cols = len(sheet["headers"])
        # Row 1: instruction banner spanning all columns
        ws.cell(row=1, column=1, value=sheet["instruction"]).font = instruction_font
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = 45
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        # Row 2: header
        for col_idx, header in enumerate(sheet["headers"], start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 40

        # Rows 3+: the engagement's current data for this section.
        rows = data.get(sheet["key"], []) or []
        for r_offset, row in enumerate(rows):
            r_idx = 3 + r_offset
            for col_idx, rk in enumerate(sheet["row_keys"], start=1):
                cell = ws.cell(row=r_idx, column=col_idx, value=row.get(rk, ""))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r_idx].height = 30

        # Keep the frequency dropdown so edited rows stay valid on re-upload.
        if sheet["key"] == "existing_reports":
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(_PREWORK_FREQ_OPTIONS) + '"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add("C3:C200")

        # Keep the Type dropdown so edited rows stay valid on re-upload.
        if sheet["key"] == "question_bank":
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(QUESTION_TYPE_OPTIONS) + '"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add("B3:B200")

    # Hidden metadata sheet — same version key the parser reads, so the export
    # re-uploads cleanly. `kind=export` distinguishes it from a blank template.
    meta = wb.create_sheet("_meta")
    meta["A1"] = "template_version"
    meta["B1"] = PREWORK_TEMPLATE_VERSION
    meta["A2"] = "generated_at"
    meta["B2"] = now_ts()
    meta["A3"] = "kind"
    meta["B3"] = "export"
    meta.sheet_state = "hidden"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_prework_xlsx(file_bytes):
    """Parse a filled-in pre-work workbook.

    Returns (parsed, warnings, errors, template_version).
      parsed:   dict keyed by section key with arrays of row-dicts. Empty
                rows are dropped. Seeded example rows on the Key Terms sheet
                are detected and skipped.
      warnings: non-fatal (template version mismatch, unknown frequency value,
                BO edited a standard S1 question prompt).
      errors:   fatal (missing sheets, renamed column headers, parse failure).
      template_version: string from the _meta sheet, "" if absent.

    Strictly read-only; no DB writes. Caller decides whether to apply.
    """
    from openpyxl import load_workbook

    warnings = []
    errors = []
    parsed = {sheet["key"]: [] for sheet in _PREWORK_SHEETS}
    template_version = ""

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return parsed, warnings, [f"Could not open file as .xlsx: {e}"], ""

    if "_meta" in wb.sheetnames:
        try:
            template_version = str(wb["_meta"]["B1"].value or "").strip()
        except Exception:
            template_version = ""
    if not template_version:
        warnings.append(
            "Template version not found in file. This may not have come from "
            "the app's template -- if you see import problems, download a "
            "fresh template from the upload dialog."
        )
    elif template_version != PREWORK_TEMPLATE_VERSION:
        warnings.append(
            f"Template version mismatch (file: {template_version}, current: "
            f"{PREWORK_TEMPLATE_VERSION}). Some fields may not import "
            "correctly. Download a fresh template if you see issues."
        )

    standard_s1_questions = {pair[0] for pair in _PREWORK_S1_CONTEXT}

    for sheet_cfg in _PREWORK_SHEETS:
        name = sheet_cfg["name"]
        if name not in wb.sheetnames:
            errors.append(
                f"Sheet '{name}' is missing. Don't rename or delete sheets — "
                "re-download the template if needed."
            )
            continue
        ws = wb[name]

        # Validate header row (row 2). Case-insensitive, whitespace-trimmed.
        actual = []
        for col_idx in range(1, len(sheet_cfg["headers"]) + 1):
            v = ws.cell(row=2, column=col_idx).value
            actual.append((v or "").strip() if isinstance(v, str) else str(v or "").strip())
        expected = [h.strip() for h in sheet_cfg["headers"]]
        if [h.lower() for h in actual] != [h.lower() for h in expected]:
            errors.append(
                f"Sheet '{name}' headers don't match. Expected {expected!r}, "
                f"got {actual!r}. Re-download the template if columns were changed."
            )
            continue

        # Identify the seeded-example rows (Key Terms AND Question Bank) so we
        # don't import them. Triggered by an "EXAMPLES" banner in column A; the
        # next three rows are skipped.
        skip_rows = set()
        if sheet_cfg["key"] in ("vocabulary_metrics", "question_bank"):
            for r_idx in range(3, min(ws.max_row, 12) + 1):
                v = ws.cell(row=r_idx, column=1).value
                if isinstance(v, str) and "examples" in v.lower() and (
                    "delete" in v.lower() or "replace" in v.lower()
                ):
                    skip_rows.add(r_idx)
                    skip_rows.update({r_idx + 1, r_idx + 2, r_idx + 3})
                    break

        for r_idx in range(3, ws.max_row + 1):
            if r_idx in skip_rows:
                continue
            values = []
            for col_idx in range(1, len(sheet_cfg["row_keys"]) + 1):
                v = ws.cell(row=r_idx, column=col_idx).value
                if v is None:
                    values.append("")
                elif isinstance(v, str):
                    values.append(v.strip())
                else:
                    values.append(str(v).strip())
            if not any(values):
                continue
            row_dict = dict(zip(sheet_cfg["row_keys"], values))

            if sheet_cfg["key"] == "existing_reports":
                freq = row_dict.get("frequency", "")
                if freq and freq not in _PREWORK_FREQ_OPTIONS:
                    warnings.append(
                        f"'{name}' row {r_idx}: 'How Often Used' value "
                        f"'{freq}' isn't one of {_PREWORK_FREQ_OPTIONS}. "
                        "Imported as-is; you can fix it in the app."
                    )

            if sheet_cfg["key"] == "business_context":
                q = row_dict.get("question", "")
                if q and q not in standard_s1_questions:
                    warnings.append(
                        f"'{name}' row {r_idx}: question text differs from "
                        "the standard prompt. The response will still import."
                    )

            parsed[sheet_cfg["key"]].append(row_dict)

    return parsed, warnings, errors, template_version


def _apply_prework_atomic(eid, sections_to_apply, parsed):
    """Atomically replace the chosen S1/S2 section columns on an engagement.

    sections_to_apply: iterable of section keys (e.g. {'business_context',
                       'question_bank'}). Anything not listed is left alone.
    parsed:            dict from _parse_prework_xlsx, keyed by section key.

    Writes all chosen columns in a single UPDATE so a partial failure can't
    leave the engagement half-applied. Honors If-Match optimistic lock the
    same way save_session does. Returns the new updated_at.
    """
    _check_optimistic_lock(eid)

    section_to_session = {s["key"]: s["session"] for s in _PREWORK_SHEETS}
    valid_keys = set(section_to_session.keys())
    keys = [k for k in sections_to_apply if k in valid_keys]
    if not keys:
        return _read_updated_at(eid)

    ts = now_ts()
    params = {"eid": eid, "ts": ts}
    set_parts = []
    for k in keys:
        set_parts.append(f"{k} = :{k}")
        params[k] = json.dumps(parsed.get(k, []))

    # Unlock the next tab: advance current_session to one past the highest
    # touched session, but never go backwards.
    max_session = max(section_to_session[k] for k in keys)
    set_parts.append(f"current_session = GREATEST(current_session, {max_session + 1})")
    set_parts.append(
        "status = CASE WHEN status IN ('complete', 'ready_for_pilot') "
        "THEN status ELSE 'in_progress' END"
    )
    set_parts.append("updated_at = :ts")

    sql_run(
        f"UPDATE {TABLE} SET {', '.join(set_parts)} WHERE engagement_id = :eid",
        params,
    )
    return _read_updated_at(eid)


@app.route("/api/template/business-owner-prework.xlsx", methods=["GET"])
def download_prework_template():
    """Stream the BO pre-work .xlsx template. Generated fresh per request so
    we never ship a stale binary."""
    try:
        buf = _build_prework_template()
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to build template: {e}"}), 500
    resp = send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="genie-discovery-bo-prework.xlsx",
    )
    # The template lives at a fixed URL; without this the browser can serve a
    # stale cached copy after the template format changes.
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/api/engagements/<eid>/parse-prework", methods=["POST"])
def parse_prework(eid):
    """Parse an uploaded pre-work .xlsx and return a preview WITHOUT mutating.

    multipart/form-data, field name `file`. Returns:
      { template_version, warnings: [...], errors: [...],
        preview: { section_key: [row_dict, ...] } }

    If `errors` is non-empty the client must show them and block apply.
    `warnings` are non-fatal and shown alongside the preview.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded. Use the 'file' form field."}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename in upload."}), 400
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({
            "error": f"File must be a .xlsx workbook (got '{f.filename}'). "
                     "Re-save the file as .xlsx and try again."
        }), 400
    raw = f.read()
    if len(raw) > PREWORK_MAX_BYTES:
        return jsonify({
            "error": f"File too large ({len(raw)} bytes). Max is {PREWORK_MAX_BYTES} bytes."
        }), 413
    if len(raw) < 100:
        return jsonify({"error": "File appears empty or truncated."}), 400
    # .xlsx is a zip; magic bytes catch the "wrong format saved with right extension" case
    if not raw.startswith(b"PK\x03\x04"):
        return jsonify({
            "error": "File is not a valid .xlsx (zip signature missing). "
                     "It may have been saved as .xls or .csv with the wrong extension."
        }), 400

    parsed, warnings, errors, version = _parse_prework_xlsx(raw)
    return jsonify({
        "template_version": version,
        "warnings": warnings,
        "errors": errors,
        "preview": parsed,
    })


@app.route("/api/engagements/<eid>/apply-prework", methods=["POST"])
def apply_prework(eid):
    """Atomically apply parsed pre-work to an engagement.

    Body: { sections: [section_key, ...], data: {section_key: [row_dict, ...]} }
    The client sends back the data it parsed plus the sections the user opted
    to apply (per-section checkboxes in the preview). Honors If-Match for
    optimistic locking; returns 409 on stale.
    """
    payload = request.get_json(silent=True) or {}
    sections = payload.get("sections") or []
    data = payload.get("data") or {}
    if not isinstance(sections, list) or not isinstance(data, dict):
        return jsonify({"error": "Body must be {sections: [...], data: {...}}"}), 400

    valid_keys = {s["key"] for s in _PREWORK_SHEETS}
    sections_set = {s for s in sections if s in valid_keys}
    if not sections_set:
        return jsonify({"error": "No valid sections selected to apply."}), 400

    # Defensive normalization: only keep recognized row keys per section so a
    # crafted payload can't smuggle extra columns into the DB write. Strip and
    # coerce-to-str so we never serialize an unexpected type.
    normalized = {}
    for s_cfg in _PREWORK_SHEETS:
        k = s_cfg["key"]
        if k not in sections_set:
            continue
        rows = data.get(k) or []
        if not isinstance(rows, list):
            return jsonify({"error": f"data.{k} must be a list."}), 400
        normalized[k] = []
        for r in rows:
            if not isinstance(r, dict):
                continue
            clean = {rk: str(r.get(rk) or "").strip() for rk in s_cfg["row_keys"]}
            if any(clean.values()):
                normalized[k].append(clean)

    try:
        ts = _apply_prework_atomic(eid, sections_set, normalized)
    except StaleEngagementError as e:
        return jsonify({
            "error": "stale",
            "current_updated_at": e.current_updated_at,
            "message": "This engagement was updated by another user. Refresh to continue.",
        }), 409
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    return jsonify({"success": True, "updated_at": ts, "applied": sorted(sections_set)})


def _build_export_csv(selected_keys, data):
    """Build a single flat CSV of the selected S1/S2 sections' current data.

    One row per item, with a leading `section` column naming the source sheet
    and a union of the selected sheets' row_keys as the remaining columns
    (blank where a column doesn't apply to that section). Returned as a
    UTF-8 BytesIO. This flat shape is what Genie Code ingests (it rejects
    .xlsx); tune the layout during real-upload verification.
    """
    import csv
    from io import StringIO

    sheets = [s for s in _PREWORK_SHEETS if s["key"] in selected_keys]
    # Union of row_keys in canonical (sheet, then within-sheet) order.
    key_order = []
    for s in sheets:
        for rk in s["row_keys"]:
            if rk not in key_order:
                key_order.append(rk)

    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(["section"] + key_order)
    for s in sheets:
        for row in data.get(s["key"], []) or []:
            writer.writerow([s["name"]] + [row.get(rk, "") for rk in key_order])

    return BytesIO(sio.getvalue().encode("utf-8"))


@app.route("/api/engagements/<eid>/export-prework", methods=["POST"])
def export_prework(eid):
    """Export selected S1/S2 sections to a populated .xlsx or .csv.

    Body: { sections: [section_key, ...], data: {section_key: [row_dict, ...]},
            format: "xlsx" | "csv" }
    The client posts the data it currently holds (WYSIWYG with the open forms).
    Read-only: builds and streams the file; no DB mutation, so no optimistic
    lock. The .xlsx matches the template shape and is re-uploadable via
    parse/apply-prework; the .csv is a flat export for Genie Code ingestion.
    """
    payload = request.get_json(silent=True) or {}
    sections = payload.get("sections") or []
    data = payload.get("data") or {}
    fmt = (payload.get("format") or "xlsx").strip().lower()
    if fmt not in ("xlsx", "csv"):
        return jsonify({"error": "format must be 'xlsx' or 'csv'."}), 400
    if not isinstance(sections, list) or not isinstance(data, dict):
        return jsonify({"error": "Body must be {sections: [...], data: {...}}"}), 400

    valid_keys = {s["key"] for s in _PREWORK_SHEETS}
    sections_set = {s for s in sections if s in valid_keys}
    if not sections_set:
        return jsonify({"error": "No valid sections selected to export."}), 400

    # Same defensive normalization as apply-prework: keep only recognized row
    # keys per section, coerce to trimmed strings, drop fully-empty rows.
    normalized = {}
    for s_cfg in _PREWORK_SHEETS:
        k = s_cfg["key"]
        if k not in sections_set:
            continue
        rows = data.get(k) or []
        if not isinstance(rows, list):
            return jsonify({"error": f"data.{k} must be a list."}), 400
        normalized[k] = []
        for r in rows:
            if not isinstance(r, dict):
                continue
            clean = {rk: str(r.get(rk) or "").strip() for rk in s_cfg["row_keys"]}
            if any(clean.values()):
                normalized[k].append(clean)

    try:
        if fmt == "csv":
            buf = _build_export_csv(sections_set, normalized)
            mimetype = "text/csv"
            download_name = "genie-discovery-export.csv"
        else:
            buf = _build_prework_export(sections_set, normalized)
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            download_name = "genie-discovery-export.xlsx"
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to build export: {e}"}), 500
    return send_file(buf, mimetype=mimetype, as_attachment=True, download_name=download_name)


# ---------------------------------------------------------------------------
# API: COE Approval
# ---------------------------------------------------------------------------

@app.route("/api/engagements/<eid>/coe-approve", methods=["PUT"])
def coe_approve(eid):
    """Set COE approval status. Server-side enforced: caller must be in COE_GROUP.

    Engagement-existence / stakeholder access is already checked by the
    before_request gate. This handler additionally requires the caller to be a
    COE member, so the status cannot be flipped by a direct API call from an
    analyst or BO.
    """
    user_w = _user_workspace_client()
    if not user_w:
        return jsonify({"error": "reauth_required"}), 401
    if not _user_is_coe_member(user_w):
        return jsonify({
            "error": f"Only members of the '{COE_GROUP}' group can approve engagements.",
        }), 403
    data = request.json
    status = data.get("status", "")
    notes = data.get("notes", "")
    reviewer = get_current_user()
    ts = now_ts()
    # The S4 COE approval owns the engagement's terminal state — this is the
    # app's final gate (piloting happens outside the app):
    #   approved          -> 'ready_for_pilot' (home-list chip: "Ready for Pilot")
    #   changes_requested -> back to 'in_progress' (re-enters the review flow)
    # anything else (e.g. pending) leaves the engagement in progress.
    if status == "approved":
        extra_set = ", status = 'ready_for_pilot'"
    elif status == "changes_requested":
        extra_set = ", status = 'in_progress'"
    else:
        extra_set = ""
    sql_run(
        f"UPDATE {TABLE} SET "
        f"coe_approval_status = :status, coe_approval_notes = :notes, "
        f"coe_reviewer_email = :reviewer, updated_at = :ts{extra_set} "
        f"WHERE engagement_id = :eid",
        {"eid": eid, "status": status, "notes": notes, "reviewer": reviewer, "ts": ts},
    )
    return jsonify({"success": True, "updated_at": ts})


# ---------------------------------------------------------------------------
# API: Unity Catalog metadata
#
# ALL endpoints below run under OBO so analysts only see catalogs/schemas/
# tables/columns/joins that their UC grants permit. The service principal is
# deliberately NOT used as a fallback, since that would leak metadata the user
# cannot actually query. If the forwarded token is missing or lacks the
# catalog.* user scopes, the endpoint returns 401 reauth_required.
# ---------------------------------------------------------------------------

def _require_obo():
    """Shared helper for UC endpoints: return a user-OBO client or a 401 response."""
    user_w = _user_workspace_client()
    if not user_w:
        return None, (jsonify({"error": "reauth_required"}), 401)
    return user_w, None


@app.route("/api/uc/catalogs")
def uc_catalogs():
    user_w, err = _require_obo()
    if err:
        return err
    try:
        cats = list(user_w.catalogs.list())
    except Exception as e:
        # Surface the failure instead of swallowing it -- a silent empty list
        # leaves the user staring at an empty picker with no idea why. Auth
        # errors translate to a reauth_required code the frontend already
        # knows how to render; everything else gets a generic error message.
        msg = f"{type(e).__name__}: {e}"
        print(f"[/api/uc/catalogs] {msg}", flush=True)
        status = 401 if "401" in msg or "PERMISSION_DENIED" in msg or "unauthorized" in msg.lower() else 502
        code = "reauth_required" if status == 401 else "catalogs_list_failed"
        return jsonify({"error": code, "message": msg}), status
    names = [c.name for c in cats if c.name and not c.name.startswith("__")]
    names.sort(key=str.lower)
    return jsonify(names)


@app.route("/api/uc/schemas")
def uc_schemas():
    user_w, err = _require_obo()
    if err:
        return err
    catalog = request.args.get("catalog", "")
    if not catalog:
        return jsonify([])
    try:
        schemas = list(user_w.schemas.list(catalog_name=catalog))
    except Exception as e:
        print(f"[/api/uc/schemas] {type(e).__name__}: {e}", flush=True)
        return jsonify([])
    names = [s.name for s in schemas if s.name and s.name != "information_schema"]
    names.sort(key=str.lower)
    return jsonify(names)


@app.route("/api/uc/tables")
def uc_tables():
    user_w, err = _require_obo()
    if err:
        return err
    catalog = request.args.get("catalog", "")
    schema = request.args.get("schema", "")
    if not catalog or not schema:
        return jsonify([])
    try:
        tables = list(user_w.tables.list(catalog_name=catalog, schema_name=schema))
    except Exception as e:
        print(f"[/api/uc/tables] {type(e).__name__}: {e}", flush=True)
        return jsonify([])
    names = [t.name for t in tables if t.name]
    names.sort(key=str.lower)
    return jsonify(names)


@app.route("/api/uc/columns")
def uc_columns():
    user_w, err = _require_obo()
    if err:
        return err
    catalog = request.args.get("catalog", "")
    schema = request.args.get("schema", "")
    table = request.args.get("table", "")
    if not catalog or not schema or not table:
        return jsonify([])
    try:
        info = user_w.tables.get(f"{catalog}.{schema}.{table}")
    except Exception as e:
        print(f"[/api/uc/columns] {type(e).__name__}: {e}", flush=True)
        return jsonify([])
    cols = getattr(info, "columns", None) or []
    return jsonify([
        {"name": c.name, "type": str(c.type_text or c.type_name or "")}
        for c in cols if getattr(c, "name", None)
    ])


@app.route("/api/uc/joins")
def uc_joins():
    """Auto-detect PK/FK relationships between selected tables via OBO tables.get()."""
    user_w, err = _require_obo()
    if err:
        return err
    tables = request.args.getlist("table")
    if len(tables) < 2:
        return jsonify([])
    in_scope = set(tables)
    results = []
    seen = set()
    for tbl in tables:
        if tbl.count(".") != 2:
            continue
        try:
            info = user_w.tables.get(tbl)
        except Exception as e:
            print(f"[/api/uc/joins] tables.get({tbl}) {type(e).__name__}: {e}", flush=True)
            continue
        for c in getattr(info, "table_constraints", None) or []:
            fk = getattr(c, "foreign_key_constraint", None)
            if not fk:
                continue
            parent = getattr(fk, "parent_table", None)
            if not parent or parent not in in_scope:
                continue
            child_cols = list(getattr(fk, "child_columns", []) or [])
            parent_cols = list(getattr(fk, "parent_columns", []) or [])
            key = (tbl, tuple(child_cols), parent, tuple(parent_cols))
            if key in seen:
                continue
            seen.add(key)
            short_child = tbl.split(".")[-1]
            short_parent = parent.split(".")[-1]
            keys_str = " AND ".join(
                f"{short_child}.{cc} = {short_parent}.{pc}"
                for cc, pc in zip(child_cols, parent_cols)
            ) or f"{short_child} = {short_parent}"
            results.append({
                "table": f"{short_child} -> {short_parent}",
                "keys": keys_str,
            })
    return jsonify(results)


@app.route("/api/uc/metric-views")
def uc_metric_views():
    """Detect existing metric views in a catalog.schema.

    Uses the UC tables REST list endpoint (not the SDK's tables.list)
    because databricks-sdk 0.44.0 returns METRIC_VIEW as TableType.MANAGED
    in TableInfo -- the enum value was added in a later SDK version. REST
    response carries table_type as a string, so the filter works.
    """
    user_w, err = _require_obo()
    if err:
        return err
    catalog_schema = request.args.get("catalog_schema", "")
    if not catalog_schema or catalog_schema.count(".") != 1:
        return jsonify([])
    cat, sch = catalog_schema.split(".", 1)
    try:
        r = requests.get(
            f"{user_w.config.host.rstrip('/')}/api/2.1/unity-catalog/tables",
            params={"catalog_name": cat, "schema_name": sch, "max_results": 200},
            headers={"Authorization": f"Bearer {user_w.config.token}"},
            timeout=15,
        )
        if not r.ok:
            print(f"[/api/uc/metric-views] {r.status_code} {r.text[:200]}", flush=True)
            return jsonify([])
    except Exception as e:
        print(f"[/api/uc/metric-views] {type(e).__name__}: {e}", flush=True)
        return jsonify([])
    results = [
        f"{cat}.{sch}.{t['name']}"
        for t in (r.json().get("tables") or [])
        if t.get("table_type") == "METRIC_VIEW" and t.get("name")
    ]
    return jsonify(results)


@app.route("/api/uc/table-type")
def uc_table_type():
    """Return UC's authoritative table_type for a single FQN.

    Used by S3's Data Sources panel: when the analyst picks a name via
    UCTablePicker, we don't know if it's a managed table, a view, or a
    metric view -- the picker dropdown lists all of them by name. This
    lookup tells us how to categorize on Add.
    """
    user_w, err = _require_obo()
    if err:
        return err
    fqn = request.args.get("fqn", "")
    if not fqn or fqn.count(".") != 2:
        return jsonify({"error": "fqn (3-part) is required"}), 400
    try:
        r = requests.get(
            f"{user_w.config.host.rstrip('/')}/api/2.1/unity-catalog/tables/{fqn}",
            headers={"Authorization": f"Bearer {user_w.config.token}"},
            timeout=15,
        )
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 502
    if not r.ok:
        return jsonify({"error": f"{r.status_code}: {r.text[:200]}"}), r.status_code
    body = r.json()
    return jsonify({
        "fqn": fqn,
        "table_type": body.get("table_type") or "MANAGED",
        "comment": body.get("comment") or "",
    })


def _list_mvs_in_schema(host, hdrs, cat, sch):
    """List MV candidate FQNs in a single (catalog, schema) via REST.

    Returns (fqns, error_message) -- error is None on success. The caller
    surfaces errors in the response so the UI can distinguish "0 MVs in
    scope" from "couldn't read this schema."
    """
    try:
        r = requests.get(
            f"{host}/api/2.1/unity-catalog/tables",
            params={"catalog_name": cat, "schema_name": sch, "max_results": 200},
            headers=hdrs, timeout=15,
        )
    except Exception as e:
        return [], f"{type(e).__name__}: {e}"
    if not r.ok:
        return [], f"HTTP {r.status_code}: {r.text[:200]}"
    fqns = [
        f"{cat}.{sch}.{t['name']}"
        for t in (r.json().get("tables") or [])
        if t.get("table_type") == "METRIC_VIEW" and t.get("name")
    ]
    return fqns, None


def _fetch_mv_dependencies(host, hdrs, fqn):
    """Return ({fqn metadata + dependencies set}, error_message). The
    caller filters by the intersect with picked tables. Per-MV REST calls
    are parallelized in the discovery endpoint via ThreadPoolExecutor."""
    try:
        r = requests.get(
            f"{host}/api/2.1/unity-catalog/tables/{fqn}",
            headers=hdrs, timeout=15,
        )
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"
    if not r.ok:
        return None, f"HTTP {r.status_code}"
    body = r.json()
    deps_raw = (body.get("view_dependencies") or {}).get("dependencies") or []
    dep_fqns = set()
    for d in deps_raw:
        tbl = (d or {}).get("table") or {}
        fn = tbl.get("table_full_name")
        if fn:
            dep_fqns.add(fn)
    return {
        "body": body,
        "dep_fqns": dep_fqns,
    }, None


def _broad_mv_scan(user_w, warehouse_id):
    """Use system.information_schema.tables to find every metric view
    visible to the user, regardless of catalog/schema. Catches MVs in
    personal catalogs that depend on shared source tables -- a common
    DSA pattern that the per-schema scan misses.

    Returns (fqns, error). On any failure (permission denied, warehouse
    cold, system table not enabled) returns ([], err) and the caller falls
    back to per-schema scan.
    """
    if not warehouse_id:
        return [], "no warehouse_id provided"
    try:
        resp = user_w.statement_execution.execute_statement(
            warehouse_id=warehouse_id,
            statement="""
                SELECT table_catalog || '.' || table_schema || '.' || table_name AS fqn
                FROM system.information_schema.tables
                WHERE table_type = 'METRIC_VIEW'
            """,
            wait_timeout="30s",
        )
    except Exception as e:
        return [], f"{type(e).__name__}: {e}"
    state = ""
    if resp.status and resp.status.state:
        state = str(resp.status.state.value if hasattr(resp.status.state, "value")
                    else resp.status.state)
    if state != "SUCCEEDED" or not resp.result:
        msg = ""
        if resp.status and resp.status.error:
            msg = (resp.status.error.message or "")[:200]
        return [], f"system.information_schema query failed ({state}): {msg}"
    rows = resp.result.data_array or []
    return [r[0] for r in rows if r and r[0]], None


@app.route("/api/uc/metric-views-for-tables")
def uc_metric_views_for_tables():
    """Find Metric Views that depend on any of the given source tables.

    Used by S3's data-sources-first flow: the analyst picks tables, the app
    surfaces existing MVs that already use those tables so the analyst can
    reuse them instead of re-authoring measures from scratch.

    Query params:
        fqns: comma-separated catalog.schema.table list (required)
        warehouse_id: optional. If provided, the discovery also queries
            system.information_schema.tables for a broader scan that catches
            MVs in catalogs/schemas different from the picked tables (common
            DSA pattern where MVs live in personal catalogs but reference
            shared source tables). If absent or the broad scan fails, falls
            back to per-(catalog,schema) scan only.

    Response shape:
        {
            "metric_views": [{fqn, catalog, schema, name, comment, owner,
                              updated_at, dependencies}],
            "errors": ["<schema>: <message>", ...],  -- non-empty when one or
                more candidate enumeration calls failed; lets the UI
                distinguish "0 matches" from "couldn't search."
            "warnings": ["..."]  -- non-fatal notes (e.g. broad scan unavailable,
                falling back to schema-only).
            "scope": {
                "broad": bool  -- true iff system.information_schema was used
            }
        }
    """
    user_w, err = _require_obo()
    if err:
        return err
    raw = request.args.get("fqns", "")
    warehouse_id = (request.args.get("warehouse_id") or "").strip()
    picked = {
        t.strip() for t in raw.split(",")
        if t.strip() and t.count(".") == 2
    }
    if not picked:
        return jsonify({"metric_views": [], "errors": [], "warnings": [], "scope": {"broad": False}})

    host = user_w.config.host.rstrip("/")
    hdrs = {"Authorization": f"Bearer {user_w.config.token}"}
    errors = []
    warnings = []
    candidates = set()
    broad_used = False

    # Step 1a: try the broad scan first if a warehouse is available. One SQL
    # query finds every MV visible to the user. If it fails (system table
    # disabled, no warehouse access, etc.) we silently fall back to the
    # narrower per-schema scan -- but the UI gets a warning so the user
    # knows discovery was scoped.
    if warehouse_id:
        broad_fqns, broad_err = _broad_mv_scan(user_w, warehouse_id)
        if broad_err:
            warnings.append(
                f"Broad MV scan via system.information_schema failed: {broad_err}. "
                "Falling back to schema-scoped scan (MVs in different catalogs/"
                "schemas from your picked tables won't be found)."
            )
        else:
            candidates.update(broad_fqns)
            broad_used = True

    # Step 1b: schema-scoped scan as fallback / supplement. Parallelize
    # across the (catalog, schema) pairs of picked tables -- usually 1-3
    # schemas, but parallelism shaves the round-trip cost regardless.
    schemas = {".".join(t.split(".")[:2]) for t in picked}
    if not broad_used:
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = {
                pool.submit(_list_mvs_in_schema, host, hdrs, *cs.split(".", 1)): cs
                for cs in schemas
            }
            for fut in futures:
                cs = futures[fut]
                fqns, scan_err = fut.result()
                if scan_err:
                    errors.append(f"{cs}: {scan_err}")
                else:
                    candidates.update(fqns)

    if not candidates:
        return jsonify({
            "metric_views": [],
            "errors": errors,
            "warnings": warnings,
            "scope": {"broad": broad_used},
        })

    # Step 2: REST-fetch view_dependencies for every candidate in parallel.
    # The LIST response above returns view_dependencies: null, so we need
    # per-MV GETs to read the dependency list. Parallelism here is the
    # biggest win -- a schema with 30 MVs goes from sequential ~6s to a
    # bounded ~1s. Failed fetches are logged-and-skipped (e.g. private MVs
    # the user lacks grants on).
    results = []
    with ThreadPoolExecutor(max_workers=16) as pool:
        futures = {
            pool.submit(_fetch_mv_dependencies, host, hdrs, fqn): fqn
            for fqn in sorted(candidates)
        }
        for fut in futures:
            fqn = futures[fut]
            data, dep_err = fut.result()
            if dep_err:
                print(f"[/api/uc/metric-views-for-tables] {fqn}: {dep_err}", flush=True)
                continue
            if not (data["dep_fqns"] & picked):
                continue
            body = data["body"]
            parts = fqn.split(".")
            results.append({
                "fqn": fqn,
                "catalog": parts[0],
                "schema": parts[1],
                "name": parts[2],
                "comment": body.get("comment", "") or "",
                "owner": body.get("owner", "") or "",
                "updated_at": body.get("updated_at"),
                "dependencies": sorted(data["dep_fqns"]),
            })

    results.sort(key=lambda x: x["fqn"])
    return jsonify({
        "metric_views": results,
        "errors": errors,
        "warnings": warnings,
        "scope": {"broad": broad_used},
    })


@app.route("/api/uc/metric-view-details")
def uc_metric_view_details():
    """Return structured dimensions + measures for a Metric View.

    Used by S3's data-sources panel to render "what does this MV cover?" --
    column names, display names, synonyms, comments, and measure markers --
    without needing an LLM. The deterministic part of Phase 1.

    Query params:
        fqn:          catalog.schema.metric_view_name (required)
        warehouse_id: SQL warehouse to run DESCRIBE on (required)

    Returns:
        { fqn, dimensions: [...], measures: [...] }
      where each entry is:
        { name, display_name, synonyms: [...], comment, data_type }

    Measure detection: DESCRIBE EXTENDED on a metric view returns each
    column's data_type with a trailing " measure" marker (e.g.
    "bigint measure"). We split on that marker.
    """
    user_w, err = _require_obo()
    if err:
        return err
    fqn = request.args.get("fqn", "")
    warehouse_id = request.args.get("warehouse_id", "")
    if not fqn or fqn.count(".") != 2:
        return jsonify({"error": "fqn (3-part) is required"}), 400
    if not warehouse_id:
        return jsonify({"error": "warehouse_id is required"}), 400

    parts = fqn.split(".")
    stmt = f"DESCRIBE EXTENDED `{parts[0]}`.`{parts[1]}`.`{parts[2]}`"
    try:
        resp = user_w.statement_execution.execute_statement(
            warehouse_id=warehouse_id, statement=stmt, wait_timeout="30s",
        )
    except Exception as e:
        print(f"[/api/uc/metric-view-details] {type(e).__name__}: {e}", flush=True)
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 502

    state = ""
    if resp.status and resp.status.state:
        state = str(resp.status.state.value if hasattr(resp.status.state, "value")
                    else resp.status.state)
    if state != "SUCCEEDED" or not resp.result or not resp.result.data_array:
        msg = ""
        if resp.status and resp.status.error:
            msg = (resp.status.error.message or "")[:300]
        return jsonify({"error": f"DESCRIBE failed ({state}): {msg}"}), 502

    dimensions, measures = [], []
    for row in resp.result.data_array:
        if not row:
            continue
        col_name = (row[0] or "").strip() if row[0] else ""
        # The "# Detailed Table Information" section comes after the
        # columns. Stop on the first blank or comment-banner row -- the
        # column metadata we care about all lives above it.
        if not col_name or col_name.startswith("#"):
            break
        data_type = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        comment = (row[2] or "").strip() if len(row) > 2 and row[2] else ""
        metadata_raw = (row[3] or "").strip() if len(row) > 3 and row[3] else ""
        meta = {}
        if metadata_raw:
            try:
                meta = json.loads(metadata_raw)
            except Exception:
                meta = {}
        entry = {
            "name": col_name,
            "data_type": data_type,
            "comment": comment,
            "display_name": (meta.get("display_name") or "").strip(),
            "synonyms": meta.get("synonyms") or [],
        }
        if "measure" in data_type.lower():
            measures.append(entry)
        else:
            dimensions.append(entry)
    return jsonify({"fqn": fqn, "dimensions": dimensions, "measures": measures})


# ---------------------------------------------------------------------------
# SPA catch-all
# ---------------------------------------------------------------------------

@app.route("/")
@app.route("/engagement/<path:path>")
@app.route("/view/<path:path>")
def serve_spa(path=None):
    return send_from_directory(app.static_folder, "index.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
