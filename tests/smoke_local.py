"""Local smoke test for the deterministic v2 Genie Discovery backend.

Mocks the Databricks WorkspaceClient so app.py imports (its module-level
ensure_table() no-ops), then swaps sql_exec/sql_run for an in-memory store
faithful to the app's exact query shapes, and exercises the real Flask routes
end-to-end plus the pure logic. No live workspace needed.
"""
import os, sys, re, io, json, traceback

os.environ.setdefault("CATALOG", "genie_training")
os.environ.setdefault("SCHEMA", "genie_discovery")
os.environ.setdefault("DATABRICKS_WAREHOUSE_ID", "dummy-wh")
os.environ.setdefault("COE_GROUP_NAME", "genie-coe-reviewers")
os.environ.setdefault("BO_GROUP_NAME", "genie-bo-reviewers")

APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, APP_DIR)
os.chdir(APP_DIR)

# --- Mock WorkspaceClient BEFORE importing app so `w = WorkspaceClient()` and
#     the module-level ensure_table() don't hit the network. ---
import databricks.sdk as _sdk
from databricks.sdk.service.sql import StatementState

class _FakeStatus:
    state = StatementState.SUCCEEDED
    error = None
class _FakeResp:
    status = _FakeStatus(); result = None; manifest = None; statement_id = "stmt"
class _FakeSE:
    def execute_statement(self, **kw): return _FakeResp()
    def get_statement(self, sid): return _FakeResp()
class _FakeCurrentUser:
    def me(self):
        class M: user_name = "tester@databricks.com"; groups = []
        return M()
class _FakeWC:
    def __init__(self, *a, **k):
        self.statement_execution = _FakeSE()
        self.current_user = _FakeCurrentUser()
_sdk.WorkspaceClient = _FakeWC

import app  # noqa: E402  (import after the mock is in place)

# --------------------------------------------------------------------------
# In-memory DB faithful to the app's query shapes
# --------------------------------------------------------------------------
DB = {}  # engagement_id -> row dict (all values stored as strings, like Delta)

def _in_mem_exec(query, params=None):
    params = params or {}
    q = " ".join(query.split())
    if "COUNT(*) AS cnt" in q:
        name = params.get("name")
        excl = params.get("eid")
        n = 0
        for eid, row in DB.items():
            if row.get("deleted_at"):
                continue
            if row.get("genie_space_name") == name and eid != excl:
                n += 1
        return [{"cnt": n}]
    if q.startswith("SELECT updated_at "):
        eid = params.get("eid")
        r = DB.get(eid)
        return [{"updated_at": r["updated_at"]}] if r else []
    if "WHERE engagement_id = :eid" in q:  # full-row fetch (authorize / get)
        eid = params.get("eid")
        r = DB.get(eid)
        if not r:
            return []
        out = dict(r)
        if "updated_at_str" in q:
            out["updated_at_str"] = r.get("updated_at", "")
            out["created_at_str"] = r.get("created_at", "")
        return [out]
    # list_engagements: select many, no eid filter
    if q.startswith("SELECT") and "engagement_id" in q and "WHERE engagement_id" not in q:
        return [dict(r) for r in DB.values() if not r.get("deleted_at")]
    return []

def _in_mem_run(query, params=None, *, must_succeed=False):
    params = params or {}
    q = " ".join(query.split())
    if q.startswith("INSERT INTO"):
        eid = params["eid"]
        row = {
            "engagement_id": eid,
            "genie_space_name": params.get("space_name", ""),
            "business_owner_name": params.get("bo_name", ""),
            "business_owner_email": params.get("bo_email", ""),
            "analyst_name": params.get("a_name", ""),
            "analyst_email": params.get("a_email", ""),
            "servicenow_ticket_url": params.get("sn_url", ""),
            "current_session": 1,
            "status": "draft",
            "created_at": params.get("ts", ""),
            "updated_at": params.get("ts", ""),
            "deleted_at": "",
        }
        for col in app.ALL_SECTION_COLS:
            row[col] = params.get(f"default_{col}", "")
        DB[eid] = row
        return
    if q.startswith("UPDATE"):
        eid = params.get("eid")
        row = DB.get(eid)
        if row is None:
            return
        # param-driven column assignments (col = :col), word-boundary matched so
        # "status = :status" doesn't fire inside "coe_approval_status = :status"
        for k, v in params.items():
            if k in ("eid", "ts"):
                continue
            if re.search(r"(?<![\w:])" + re.escape(k) + r" = :" + re.escape(k) + r"\b", q):
                row[k] = v
        # coe_approve binds :status/:notes/:reviewer to the coe_approval_* columns
        if "coe_approval_status = :status" in q:
            row["coe_approval_status"] = params.get("status", "")
            row["coe_approval_notes"] = params.get("notes", "")
            row["coe_reviewer_email"] = params.get("reviewer", "")
        if "updated_at = :ts" in q and "ts" in params:
            row["updated_at"] = params["ts"]
        if "deleted_at = :ts" in q and "ts" in params:
            row["deleted_at"] = params["ts"]
        # current_session = GREATEST(current_session, N)
        m = re.search(r"current_session = GREATEST\(current_session, (\d+)\)", q)
        if m:
            row["current_session"] = max(int(row.get("current_session") or 1), int(m.group(1)))
        # status transitions (order matters: literals before CASE)
        if "coe_approval_status = 'ready_for_review'" in q:
            row["coe_approval_status"] = "ready_for_review"
        if "status = 'ready_for_pilot'" in q:
            row["status"] = "ready_for_pilot"
        elif re.search(r"(?<!approval_)status = 'in_progress'", q):
            row["status"] = "in_progress"
        elif "status = CASE WHEN status IN ('complete', 'ready_for_pilot')" in q:
            cur = row.get("status")
            row["status"] = cur if cur in ("complete", "ready_for_pilot") else "in_progress"
        return
    return

app.sql_exec = _in_mem_exec
app.sql_run = _in_mem_run
app._check_optimistic_lock = lambda eid: None  # not under test; avoid token friction

# --------------------------------------------------------------------------
# Test runner
# --------------------------------------------------------------------------
PASS, FAIL = [], []
def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(("  ok  " if cond else "FAIL  ") + name + (f"  — {detail}" if detail and not cond else ""))

client = app.app.test_client()
H = {"X-Forwarded-Email": "tester@databricks.com"}

print("\n== A. Session model ==")
check("4 sessions", set(app.SESSION_COLS) == {1, 2, 3, 4}, str(list(app.SESSION_COLS)))
check("S3 has data_plan + plan_warehouse_id",
      "data_plan" in app.SESSION_COLS[3] and "plan_warehouse_id" in app.SESSION_COLS[3])
check("S2 has question_bank + vocabulary_metrics",
      set(app.SESSION_COLS[2]) == {"question_bank", "vocabulary_metrics"})
check("S4 is COE-only", set(app.SESSION_COLS[4]) == {"coe_approval_status", "coe_approval_notes", "coe_reviewer_email"})
removed_cols = {"benchmark_questions", "plan_narrative", "metric_view_yaml", "acknowledgments",
                "prod_approval_status", "term_classifications", "sql_expressions",
                "example_queries", "clarifying_questions", "auto_summary", "analyst_commentary"}
check("removed columns gone from ALL_SECTION_COLS", not (removed_cols & set(app.ALL_SECTION_COLS)),
      str(removed_cols & set(app.ALL_SECTION_COLS)))
check("LAST_SESSION == 4", app.LAST_SESSION == 4)
check("OBJECT_COLS empty", app.OBJECT_COLS == set())

print("\n== B. Route map (removed gone / kept present) ==")
rules = {r.rule for r in app.app.url_map.iter_rules()}
def has(sub): return any(sub in r for r in rules)
for gone in ["/generate-plan", "/draft-benchmarks", "/draft-benchmark-sql", "/draft-benchmark-summary",
             "/draft-metric-view-yaml", "/mv-prompt-preview", "/create-metric-view", "/run-benchmark-sql",
             "/push-to-genie", "/prod-approve", "/acknowledge", "/space-access", "/auto-summary",
             "/benchmarks/bo-approved", "/jobs/start", "/jobs/", "/sessions/5", "/sessions/6", "/sessions/7"]:
    check(f"removed route absent: {gone}", not has(gone))
for kept in ["/sessions/1", "/sessions/2", "/sessions/3", "/sessions/4", "/coe-approve",
             "/request-review", "/export-prework", "/parse-prework", "/apply-prework",
             "/uc/catalogs", "/uc/metric-views-for-tables", "/warehouses"]:
    check(f"kept route present: {kept}", has(kept))

print("\n== C. parse_row tolerates an existing (pre-strip) engagement row ==")
legacy_row = {
    "engagement_id": "e1", "genie_space_name": "Legacy", "status": "complete",
    "current_session": 7, "updated_at": "2026-06-01 00:00:00",
    # new-model columns
    "business_context": json.dumps([{"question": "q", "response": "r"}]),
    "question_bank": json.dumps([{"question_text": "How many ED visits?", "type": "Benchmark", "decision_it_drives": "x"}]),
    "vocabulary_metrics": "[]", "data_plan": json.dumps([{"table_or_view": "cat.sch.t"}]),
    "plan_warehouse_id": "wh1", "global_filter": "voided='N'", "text_instructions": "[]",
    "data_gaps": "[]", "scope_boundaries": "[]",
    "coe_approval_status": "approved", "coe_approval_notes": "", "coe_reviewer_email": "a@b.co",
    # DORMANT legacy columns that still physically exist in the Delta table
    "benchmark_questions": json.dumps([{"question": "old"}]),
    "plan_narrative": "old plan", "metric_view_yaml": "yaml: x", "acknowledgments": "{}",
    "prod_approval_status": "approved", "term_classifications": "[]", "sql_expressions": "[]",
}
parsed = app.parse_row(legacy_row)
check("parse_row builds exactly sessions 1-4", set(parsed["sessions"].keys()) == {"1", "2", "3", "4"},
      str(set(parsed["sessions"].keys())))
check("S2 question keeps type", parsed["sessions"]["2"]["question_bank"][0].get("type") == "Benchmark")
check("S3 data_plan decoded", parsed["sessions"]["3"]["data_plan"][0]["table_or_view"] == "cat.sch.t")
check("dormant legacy col preserved on eng but not in a session",
      parsed.get("benchmark_questions") is not None
      and all("benchmark_questions" not in s for s in parsed["sessions"].values()))

print("\n== D. _bo_can_access whitelist ==")
check("BO can PUT sessions/1", app._bo_can_access("PUT", "/sessions/1"))
check("BO can PUT sessions/2", app._bo_can_access("PUT", "/sessions/2"))
check("BO cannot PUT sessions/3", not app._bo_can_access("PUT", "/sessions/3"))
check("BO cannot PUT sessions/4", not app._bo_can_access("PUT", "/sessions/4"))
check("BO cannot POST coe-approve", not app._bo_can_access("PUT", "/coe-approve"))
check("benchmarks/bo-approved no longer whitelisted", not app._bo_can_access("PATCH", "/benchmarks/bo-approved"))
check("BO can GET anything", app._bo_can_access("GET", "/anything"))
check("BO can export", app._bo_can_access("POST", "/export-prework"))

print("\n== E. Pre-work template build + parse round-trip (Type column) ==")
tmpl = app._build_prework_template().read()
from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(tmpl))
check("template has Question Bank sheet", "S2 Question Bank" in wb.sheetnames)
qb = wb["S2 Question Bank"]
hdrs = [qb.cell(row=2, column=c).value for c in range(1, 4)]
check("Question Bank headers include Type", hdrs == ["Question", "Type", "Decision It Drives"], str(hdrs))
# fill a question row (row 3) and re-parse
qb.cell(row=3, column=1, value="What was ED volume by site?")
qb.cell(row=3, column=2, value="Benchmark")
qb.cell(row=3, column=3, value="staffing")
buf = io.BytesIO(); wb.save(buf)
parsed, pwarn, perr, pver = app._parse_prework_xlsx(buf.getvalue())
check("parse returns template_version 2.0", pver == "2.0", str(pver))
check("parse has no fatal errors", not perr, str(perr))
pqb = parsed.get("question_bank", [])
check("parsed question has text+type", pqb and pqb[0].get("question_text") == "What was ED volume by site?" and pqb[0].get("type") == "Benchmark", str(pqb))

print("\n== F. Export .xlsx round-trip ==")
data = {"question_bank": [{"question_text": "Q1", "type": "Testing", "decision_it_drives": "d"}]}
xbuf = app._build_prework_export({"question_bank"}, data)
wx = load_workbook(io.BytesIO(xbuf.read()))
check("export xlsx has Question Bank sheet", "S2 Question Bank" in wx.sheetnames)
qx = wx["S2 Question Bank"]
check("exported row carries type", qx.cell(row=3, column=2).value == "Testing",
      str(qx.cell(row=3, column=2).value))

print("\n== G. Export .csv structure ==")
csv_bytes = app._build_export_csv({"question_bank", "vocabulary_metrics"}, {
    "vocabulary_metrics": [{"business_term": "ED", "what_they_mean": "emergency dept", "synonyms": "ER"}],
    "question_bank": [{"question_text": "Q1", "type": "Benchmark", "decision_it_drives": "d"}],
}).read().decode("utf-8")
lines = [l for l in csv_bytes.splitlines() if l.strip()]
check("csv header starts with section", lines[0].startswith("section,"), lines[0])
check("csv header includes type", "type" in lines[0].split(","), lines[0])
check("csv has a section-tagged question row", any("S2 Question Bank" in l and "Benchmark" in l for l in lines), csv_bytes)
check("csv has a key-terms row", any("S2 Key Terms" in l for l in lines))

print("\n== H. End-to-end HTTP: create -> save S1-S4 -> get round-trip ==")
r = client.post("/api/engagements", headers=H, json={
    "genie_space_name": "Smoke Test Space", "business_owner_name": "BO",
    "business_owner_email": "bo@x.co", "analyst_name": "AN"})
check("create 201", r.status_code == 201, f"{r.status_code} {r.get_data(as_text=True)[:200]}")
eid = r.get_json()["engagement_id"]
r1 = client.put(f"/api/engagements/{eid}/sessions/1", headers=H,
                json={"business_context": [{"question": "q", "response": "r"}], "pain_points": [], "existing_reports": []})
check("save S1 200", r1.status_code == 200, r1.get_data(as_text=True)[:200])
r2 = client.put(f"/api/engagements/{eid}/sessions/2", headers=H,
                json={"vocabulary_metrics": [{"business_term": "ED"}],
                      "question_bank": [{"question_text": "How many ED visits?", "type": "Benchmark", "decision_it_drives": "d"}]})
check("save S2 200", r2.status_code == 200)
r3 = client.put(f"/api/engagements/{eid}/sessions/3", headers=H,
                json={"data_plan": [{"table_or_view": "c.s.t", "type": "Table", "include_in_space": "Yes", "notes": ""}],
                      "plan_warehouse_id": "wh9", "global_filter": "dept='ED'",
                      "text_instructions": [], "data_gaps": [], "scope_boundaries": []})
check("save S3 200", r3.status_code == 200)
g = client.get(f"/api/engagements/{eid}", headers=H).get_json()
check("round-trip S2 question type", g["sessions"]["2"]["question_bank"][0]["type"] == "Benchmark")
check("round-trip S3 global_filter", g["sessions"]["3"]["global_filter"] == "dept='ED'")
check("round-trip S3 data_plan", g["sessions"]["3"]["data_plan"][0]["table_or_view"] == "c.s.t")
check("current_session advanced", DB[eid]["current_session"] >= 4, str(DB[eid]["current_session"]))

print("\n== I. Removed routes 404 via HTTP ==")
for path, method in [("/generate-plan", "POST"), ("/push-to-genie", "POST"), ("/prod-approve", "PUT"),
                     ("/acknowledge", "POST"), ("/sessions/5", "PUT"), ("/draft-benchmarks", "POST"),
                     ("/create-metric-view", "POST")]:
    rr = client.open(f"/api/engagements/{eid}{path}", method=method, headers=H, json={})
    # No real handler: 404, or 405 from the static catch-all (static_url_path="").
    check(f"no handler for removed {method} {path}", rr.status_code in (404, 405), str(rr.status_code))

print("\n== J. COE approval -> ready_for_pilot ==")
app._user_workspace_client = lambda: object()  # non-None OBO
app._user_is_coe_member = lambda uw: True
app._user_role = lambda uw: "coe"
ra = client.put(f"/api/engagements/{eid}/coe-approve", headers=H, json={"status": "approved", "notes": "lgtm"})
check("coe-approve 200", ra.status_code == 200, ra.get_data(as_text=True)[:200])
check("engagement status = ready_for_pilot", DB[eid]["status"] == "ready_for_pilot", DB[eid]["status"])
check("coe_approval_status = approved", DB[eid]["coe_approval_status"] == "approved")
rc = client.put(f"/api/engagements/{eid}/coe-approve", headers=H, json={"status": "changes_requested", "notes": "fix"})
check("changes_requested -> in_progress", DB[eid]["status"] == "in_progress", DB[eid]["status"])

print("\n== K. BO RBAC gate ==")
app._user_role = lambda uw: "bo"
rb3 = client.put(f"/api/engagements/{eid}/sessions/3", headers=H, json={"data_plan": []})
check("BO blocked from S3 (403)", rb3.status_code == 403, str(rb3.status_code))
rb1 = client.put(f"/api/engagements/{eid}/sessions/1", headers=H,
                 json={"business_context": [], "pain_points": [], "existing_reports": []})
check("BO allowed on S1 (200)", rb1.status_code == 200, str(rb1.status_code))
app._user_role = lambda uw: "coe"  # reset

print("\n== L. Export endpoint (xlsx + csv) via HTTP ==")
ex = client.post(f"/api/engagements/{eid}/export-prework", headers=H,
                 json={"sections": ["question_bank"], "format": "xlsx",
                       "data": {"question_bank": [{"question_text": "Q", "type": "Benchmark", "decision_it_drives": "d"}]}})
check("export xlsx 200 + mimetype", ex.status_code == 200 and "spreadsheet" in ex.headers.get("Content-Type", ""),
      f"{ex.status_code} {ex.headers.get('Content-Type')}")
ec = client.post(f"/api/engagements/{eid}/export-prework", headers=H,
                 json={"sections": ["question_bank"], "format": "csv",
                       "data": {"question_bank": [{"question_text": "Q", "type": "Benchmark", "decision_it_drives": "d"}]}})
check("export csv 200 + text/csv", ec.status_code == 200 and "csv" in ec.headers.get("Content-Type", ""),
      f"{ec.status_code} {ec.headers.get('Content-Type')}")
check("csv body has type + Benchmark", b"type" in ec.data and b"Benchmark" in ec.data)

print("\n== M. request-review flow ==")
app._user_role = lambda uw: "analyst"
rr = client.put(f"/api/engagements/{eid}/request-review", headers=H)
check("request-review 200", rr.status_code == 200, str(rr.status_code))
check("coe_approval_status ready_for_review", DB[eid]["coe_approval_status"] == "ready_for_review", DB[eid].get("coe_approval_status"))

print(f"\n=========== RESULT: {len(PASS)} passed, {len(FAIL)} failed ===========")
if FAIL:
    print("FAILURES:")
    for f in FAIL:
        print("  -", f)
    sys.exit(1)
print("ALL GREEN")
